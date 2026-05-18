"""
importers.py — Import Excel (XLSX / XLS) pour INSHOP BUILDER
=============================================================

Modèles supportés :
  - Product  (composants, périphériques, laptops)
  - Prebuilt (PC pré-montés)
  - OnlyOnePC

Comportement :  update_or_create sur le champ "name" — si un produit du même
nom existe déjà il est mis à jour, sinon il est créé.

Génération de modèles (templates) : chaque import dispose d'un endpoint
GET ?download_template=1 qui renvoie un fichier XLSX pré-rempli avec
les en-têtes et une ligne d'exemple.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Category, Product, Prebuilt, OnlyOnePC


# ── Définition des colonnes ───────────────────────────────────────────────────
# Chaque entrée : (nom_colonne_excel, champ_modèle, type_python, description, requis)

PRODUCT_COLUMNS = [
    # Général
    ('name',                 'name',                 str,   'Nom du produit',                                        True),
    ('brand',                'brand',                str,   'Marque (ex: AMD, NVIDIA, Samsung)',                     False),
    ('price',                'price',                float, 'Prix en USD (ex: 299.99)',                              False),
    ('category_code',        '__category',           str,   'Code catégorie (cpu, gpu, ram, motherboard…)',          False),
    ('stock_status',         'stock_status',         str,   'in_stock | low_stock | out_of_stock',                   False),
    ('description',          'description',          str,   'Description détaillée',                                 False),
    ('tags',                 'tags',                 str,   'Tags séparés par virgule (ex: gaming,bestseller)',      False),
    ('is_only_one',          'is_only_one',          bool,  'true/false — composant "Only One PC"',                 False),
    ('active',               'active',               bool,  'true/false — produit visible',                          False),
    # CPU
    ('spec_cpu_socket',      'spec_cpu_socket',      str,   'Socket CPU (AM5, AM4, LGA1700…)',                       False),
    ('spec_cpu_cores',       'spec_cpu_cores',       int,   'Nombre de cœurs',                                       False),
    ('spec_cpu_threads',     'spec_cpu_threads',     int,   'Nombre de threads',                                     False),
    ('spec_cpu_base_ghz',    'spec_cpu_base_ghz',    str,   'Fréquence de base (ex: 3.8)',                           False),
    ('spec_cpu_boost_ghz',   'spec_cpu_boost_ghz',   str,   'Fréquence boost (ex: 5.2)',                             False),
    ('spec_cpu_tdp',         'spec_cpu_tdp',         int,   'TDP en watts',                                          False),
    ('spec_cpu_cache',       'spec_cpu_cache',       str,   'Cache (ex: 32 Mo L3)',                                  False),
    # Carte mère
    ('spec_mb_socket',       'spec_mb_socket',       str,   'Socket carte mère',                                     False),
    ('spec_mb_form_factor',  'spec_mb_form_factor',  str,   'Format (ATX, mATX, ITX, E-ATX)',                        False),
    ('spec_mb_ram_type',     'spec_mb_ram_type',     str,   'Type RAM supporté (DDR4, DDR5…)',                       False),
    ('spec_mb_ram_slots',    'spec_mb_ram_slots',    int,   'Nombre de slots RAM',                                   False),
    ('spec_mb_m2_slots',     'spec_mb_m2_slots',     int,   'Nombre de slots M.2',                                   False),
    ('spec_mb_pcie_slots',   'spec_mb_pcie_slots',   int,   'Nombre de slots PCIe',                                  False),
    ('spec_mb_chipset',      'spec_mb_chipset',      str,   'Chipset (ex: B650, Z790)',                              False),
    # RAM
    ('spec_ram_type',        'spec_ram_type',        str,   'Type RAM (DDR5, DDR4…)',                                False),
    ('spec_ram_capacity',    'spec_ram_capacity',    str,   'Capacité (ex: 32 Go)',                                  False),
    ('spec_ram_speed',       'spec_ram_speed',       int,   'Vitesse en MHz',                                        False),
    ('spec_ram_modules',     'spec_ram_modules',     int,   'Nombre de barrettes',                                   False),
    ('spec_ram_timing',      'spec_ram_timing',      str,   'Timing (ex: CL36)',                                     False),
    # GPU
    ('spec_gpu_vram',        'spec_gpu_vram',        str,   'VRAM (ex: 16 Go)',                                      False),
    ('spec_gpu_mem_type',    'spec_gpu_mem_type',    str,   'Type mémoire (GDDR6X, GDDR6…)',                         False),
    ('spec_gpu_tdp',         'spec_gpu_tdp',         int,   'TDP GPU en watts',                                      False),
    ('spec_gpu_length_mm',   'spec_gpu_length_mm',   int,   'Longueur en mm',                                        False),
    ('spec_gpu_architecture','spec_gpu_architecture',str,   'Architecture (Ada Lovelace, RDNA 3…)',                  False),
    ('spec_gpu_outputs',     'spec_gpu_outputs',     str,   'Sorties vidéo (ex: 3x DP 1.4, 1x HDMI 2.1)',           False),
    # PSU
    ('spec_psu_wattage',     'spec_psu_wattage',     int,   'Puissance en watts',                                    False),
    ('spec_psu_efficiency',  'spec_psu_efficiency',  str,   '80+ Gold | 80+ Platinum…',                             False),
    ('spec_psu_modular',     'spec_psu_modular',     str,   'Fully Modular | Semi-Modular | Non-Modular',            False),
    # Boîtier
    ('spec_case_atx',        'spec_case_atx',        bool,  'Compatible ATX (true/false)',                           False),
    ('spec_case_matx',       'spec_case_matx',       bool,  'Compatible mATX',                                       False),
    ('spec_case_itx',        'spec_case_itx',        bool,  'Compatible ITX',                                        False),
    ('spec_case_eatx',       'spec_case_eatx',       bool,  'Compatible E-ATX',                                      False),
    ('spec_case_max_gpu_mm', 'spec_case_max_gpu_mm', int,   'GPU max en mm',                                         False),
    ('spec_case_max_cooler_mm','spec_case_max_cooler_mm',int,'Ventirad max en mm',                                   False),
    ('spec_case_drive_bays', 'spec_case_drive_bays', int,   "Nombre d'emplacements disque",                          False),
    ('spec_case_fans_included','spec_case_fans_included',int,'Ventilateurs inclus',                                  False),
    # Refroidissement
    ('spec_cool_type',       'spec_cool_type',       str,   'Air | AIO 120 | AIO 240 | AIO 360 | Custom',           False),
    ('spec_cool_tdp_rating', 'spec_cool_tdp_rating', int,   'TDP max supporté en watts',                             False),
    ('spec_cool_height_mm',  'spec_cool_height_mm',  int,   'Hauteur en mm',                                         False),
    ('spec_cool_radiator_mm','spec_cool_radiator_mm',int,   'Taille radiateur en mm',                                False),
    ('spec_cool_sock_am5',   'spec_cool_sock_am5',   bool,  'Supporte AM5',                                          False),
    ('spec_cool_sock_am4',   'spec_cool_sock_am4',   bool,  'Supporte AM4',                                          False),
    ('spec_cool_sock_lga1700','spec_cool_sock_lga1700',bool,'Supporte LGA1700',                                      False),
    ('spec_cool_sock_lga1200','spec_cool_sock_lga1200',bool,'Supporte LGA1200',                                      False),
    # Stockage
    ('spec_stor_type',       'spec_stor_type',       str,   'NVMe | SATA | HDD | SATA M.2',                         False),
    ('spec_stor_capacity',   'spec_stor_capacity',   str,   'Capacité (ex: 1 To)',                                   False),
    ('spec_stor_read',       'spec_stor_read',       str,   'Vitesse lecture (ex: 7000 Mo/s)',                       False),
    ('spec_stor_write',      'spec_stor_write',      str,   'Vitesse écriture',                                      False),
    # Laptop
    ('spec_laptop_cpu',      'spec_laptop_cpu',      str,   'Processeur laptop',                                     False),
    ('spec_laptop_gpu',      'spec_laptop_gpu',      str,   'GPU laptop',                                            False),
    ('spec_laptop_ram',      'spec_laptop_ram',      str,   'RAM laptop',                                            False),
    ('spec_laptop_storage',  'spec_laptop_storage',  str,   'Stockage laptop',                                       False),
    ('spec_laptop_display',  'spec_laptop_display',  str,   'Écran (ex: 15.6" FHD 144Hz)',                           False),
    ('spec_laptop_battery',  'spec_laptop_battery',  str,   'Batterie (ex: 72 Wh)',                                  False),
    ('spec_laptop_os',       'spec_laptop_os',       str,   'Système (Windows 11, Linux…)',                          False),
    ('spec_laptop_weight',   'spec_laptop_weight',   str,   'Poids (ex: 2.1 kg)',                                    False),
]

PREBUILT_COLUMNS = [
    ('name',                   'name',                   str,   'Nom du PC',                                True),
    ('price',                  'price',                  float, 'Prix en USD',                               False),
    ('tier',                   'tier',                   str,   'Budget | Mid-Range | High-End | Flagship',  False),
    ('badge',                  'badge',                  str,   'Badge affiché (ex: BEST-VENTE)',             False),
    ('tag_line',               'tag_line',               str,   'Sous-titre court',                           False),
    ('brand',                  'brand',                  str,   'Marque (défaut: INSHOP)',                    False),
    ('stock_status',           'stock_status',           str,   'in_stock | low_stock | out_of_stock',        False),
    ('rating',                 'rating',                 float, 'Note (0.0 – 5.0)',                           False),
    ('review_count',           'review_count',           int,   'Nombre d\'avis',                             False),
    ('gaming_perf',            'gaming_perf',            str,   'Ex: 1440p Ultra',                            False),
    ('workflow',               'workflow',               str,   'Ex: Montage vidéo & streaming',              False),
    ('cpu',                    'cpu',                    str,   'Processeur (affichage)',                     False),
    ('gpu',                    'gpu',                    str,   'Carte graphique (affichage)',                False),
    ('ram',                    'ram',                    str,   'RAM (affichage, ex: 32 Go DDR5)',            False),
    ('storage',                'storage',                str,   'Stockage (affichage)',                       False),
    ('psu',                    'psu',                    str,   'Alimentation (affichage)',                   False),
    ('case_name',              'case_name',              str,   'Boîtier (affichage)',                        False),
    ('cooling',                'cooling',                str,   'Refroidissement (affichage)',                False),
    ('tech_cpu_socket',        'tech_cpu_socket',        str,   'Socket CPU technique',                       False),
    ('tech_ram_type',          'tech_ram_type',          str,   'Type RAM technique',                         False),
    ('tech_ram_capacity',      'tech_ram_capacity',      str,   'Capacité RAM technique',                     False),
    ('tech_ram_speed',         'tech_ram_speed',         int,   'Vitesse RAM en MHz',                         False),
    ('tech_storage_type',      'tech_storage_type',      str,   'Type stockage technique',                    False),
    ('tech_storage_capacity',  'tech_storage_capacity',  str,   'Capacité stockage technique',                False),
    ('tech_gpu_vram',          'tech_gpu_vram',          str,   'VRAM GPU technique',                         False),
    ('tech_psu_wattage',       'tech_psu_wattage',       int,   'Puissance PSU en watts',                     False),
    ('tech_psu_efficiency',    'tech_psu_efficiency',    str,   'Certification 80+',                          False),
    ('active',                 'active',                 bool,  'true/false — visible',                       False),
]

ONLYONE_COLUMNS = [
    ('name',                   'name',                   str,   'Nom du PC',                                True),
    ('price',                  'price',                  float, 'Prix en USD',                               False),
    ('badge',                  'badge',                  str,   'Badge (ex: EXCLUSIF)',                       False),
    ('tag_line',               'tag_line',               str,   'Sous-titre court',                           False),
    ('stock_status',           'stock_status',           str,   'in_stock | low_stock | out_of_stock',        False),
    ('rating',                 'rating',                 float, 'Note (0.0 – 5.0)',                           False),
    ('review_count',           'review_count',           int,   'Nombre d\'avis',                             False),
    ('cpu',                    'cpu',                    str,   'Processeur',                                 False),
    ('gpu',                    'gpu',                    str,   'Carte graphique',                            False),
    ('ram',                    'ram',                    str,   'RAM',                                        False),
    ('storage',                'storage',                str,   'Stockage',                                   False),
    ('psu',                    'psu',                    str,   'Alimentation',                               False),
    ('case_name',              'case_name',              str,   'Boîtier',                                    False),
    ('cooling',                'cooling',                str,   'Refroidissement',                            False),
    ('tech_cpu_socket',        'tech_cpu_socket',        str,   'Socket CPU',                                 False),
    ('tech_ram_type',          'tech_ram_type',          str,   'Type RAM',                                   False),
    ('tech_ram_capacity',      'tech_ram_capacity',      str,   'Capacité RAM',                               False),
    ('tech_ram_speed',         'tech_ram_speed',         int,   'Vitesse RAM MHz',                            False),
    ('tech_storage_type',      'tech_storage_type',      str,   'Type stockage',                              False),
    ('tech_storage_capacity',  'tech_storage_capacity',  str,   'Capacité stockage',                          False),
    ('tech_gpu_vram',          'tech_gpu_vram',          str,   'VRAM',                                       False),
    ('tech_psu_wattage',       'tech_psu_wattage',       int,   'Puissance PSU watts',                        False),
    ('tech_psu_efficiency',    'tech_psu_efficiency',    str,   'Certification 80+',                          False),
    ('active',                 'active',                 bool,  'true/false — visible',                       False),
]

PRODUCT_EXAMPLE = {
    'name': 'AMD Ryzen 7 5800X', 'brand': 'AMD', 'price': '299.99',
    'category_code': 'cpu', 'stock_status': 'in_stock',
    'description': 'Processeur AMD 8 cœurs socket AM4',
    'tags': 'gaming,bestseller', 'is_only_one': 'false', 'active': 'true',
    'spec_cpu_socket': 'AM4', 'spec_cpu_cores': '8', 'spec_cpu_threads': '16',
    'spec_cpu_base_ghz': '3.8', 'spec_cpu_boost_ghz': '4.7',
    'spec_cpu_tdp': '105', 'spec_cpu_cache': '32 Mo L3',
}

PREBUILT_EXAMPLE = {
    'name': 'INSHOP Gaming Pro', 'price': '1499.99', 'tier': 'High-End',
    'brand': 'INSHOP', 'stock_status': 'in_stock', 'active': 'true',
    'gaming_perf': '1440p Ultra', 'workflow': 'Gaming & streaming',
    'cpu': 'AMD Ryzen 7 7700X', 'gpu': 'NVIDIA RTX 4070',
    'ram': '32 Go DDR5 6000MHz', 'storage': '1 To NVMe Gen4',
    'psu': '750W 80+ Gold', 'case_name': 'Fractal North',
    'cooling': 'AIO 240mm', 'tech_cpu_socket': 'AM5',
    'tech_ram_type': 'DDR5', 'tech_ram_capacity': '32 Go',
    'tech_ram_speed': '6000', 'tech_gpu_vram': '12 Go',
    'tech_psu_wattage': '750', 'tech_psu_efficiency': '80+ Gold',
}

ONLYONE_EXAMPLE = {
    'name': 'Only One Elite X', 'price': '2999.99',
    'badge': 'EXCLUSIF', 'tag_line': 'La machine ultime',
    'stock_status': 'in_stock', 'active': 'true',
    'cpu': 'Intel Core i9-14900K', 'gpu': 'NVIDIA RTX 4090',
    'ram': '64 Go DDR5 7200MHz', 'storage': '4 To NVMe Gen5',
    'psu': '1000W 80+ Platinum', 'tech_cpu_socket': 'LGA1700',
    'tech_gpu_vram': '24 Go', 'tech_psu_wattage': '1000',
}


# ── Utilitaires ───────────────────────────────────────────────────────────────

def _cast(value, cast_type):
    """Convertit la valeur d'une cellule dans le type Python attendu."""
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return None
    if cast_type is bool:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ('1', 'true', 'oui', 'yes', 'o', 'vrai')
    if cast_type is int:
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return None
    if cast_type is float:
        try:
            return float(str(value).strip().replace(',', '.'))
        except (ValueError, TypeError):
            return None
    # str
    v = str(value).strip()
    return v if v else None


def _parse_sheet(ws, columns):
    """
    Lit un worksheet openpyxl et retourne une liste de dicts
    {champ_modèle: valeur_castée} en fonction de la map de colonnes.
    Ignore les lignes entièrement vides.
    """
    header_row = [
        str(c.value).strip().lower() if c.value is not None else ''
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    # index du header dans la liste des colonnes connues
    col_map = {}  # header_name → (model_field, cast)
    for col_name, model_field, cast_type, _, _ in columns:
        col_map[col_name.lower()] = (model_field, cast_type)

    header_idx = {}  # model_field → colonne index
    for idx, h in enumerate(header_row):
        if h in col_map:
            field, cast = col_map[h]
            header_idx[field] = (idx, cast)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        obj = {}
        for field, (idx, cast) in header_idx.items():
            raw = row[idx] if idx < len(row) else None
            obj[field] = _cast(raw, cast)
        rows.append(obj)
    return rows


# ── Fonctions d'import ────────────────────────────────────────────────────────

def import_products(file_bytes):
    """
    Importe des produits depuis un fichier XLSX.
    Fait un update_or_create sur le champ 'name'.
    Retourne {'created': int, 'updated': int, 'errors': [str]}.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {'created': 0, 'updated': 0, 'errors': [f'Fichier invalide : {e}']}

    ws = wb.active
    category_cache = {c.code.lower(): c for c in Category.objects.all()}
    rows = _parse_sheet(ws, PRODUCT_COLUMNS)

    created, updated, errors = 0, 0, []

    for i, data in enumerate(rows, start=2):
        name = data.get('name')
        if not name:
            errors.append(f'Ligne {i} : colonne "name" manquante ou vide — ligne ignorée.')
            continue

        # Résoudre la catégorie FK
        cat_code = data.pop('__category', None)
        category = category_cache.get(cat_code.lower()) if cat_code else None

        # Supprimer les valeurs None sauf les booléens
        fields = {}
        for k, v in data.items():
            if v is not None:
                fields[k] = v
            elif isinstance(v, bool):
                fields[k] = v

        fields['category'] = category

        try:
            _, was_created = Product.objects.update_or_create(
                name=name,
                defaults=fields,
            )
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f'Ligne {i} ({name}) : {e}')

    return {'created': created, 'updated': updated, 'errors': errors}


def import_prebuilts(file_bytes):
    """Import de PC pré-montés depuis XLSX."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {'created': 0, 'updated': 0, 'errors': [f'Fichier invalide : {e}']}

    ws = wb.active
    rows = _parse_sheet(ws, PREBUILT_COLUMNS)
    created, updated, errors = 0, 0, []

    for i, data in enumerate(rows, start=2):
        name = data.get('name')
        if not name:
            errors.append(f'Ligne {i} : colonne "name" manquante — ligne ignorée.')
            continue
        fields = {k: v for k, v in data.items() if v is not None or isinstance(v, bool)}
        try:
            _, was_created = Prebuilt.objects.update_or_create(name=name, defaults=fields)
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f'Ligne {i} ({name}) : {e}')

    return {'created': created, 'updated': updated, 'errors': errors}


def import_onlyonepcs(file_bytes):
    """Import de Only One PCs depuis XLSX."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {'created': 0, 'updated': 0, 'errors': [f'Fichier invalide : {e}']}

    ws = wb.active
    rows = _parse_sheet(ws, ONLYONE_COLUMNS)
    created, updated, errors = 0, 0, []

    for i, data in enumerate(rows, start=2):
        name = data.get('name')
        if not name:
            errors.append(f'Ligne {i} : colonne "name" manquante — ligne ignorée.')
            continue
        fields = {k: v for k, v in data.items() if v is not None or isinstance(v, bool)}
        try:
            _, was_created = OnlyOnePC.objects.update_or_create(name=name, defaults=fields)
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f'Ligne {i} ({name}) : {e}')

    return {'created': created, 'updated': updated, 'errors': errors}


# ── Génération de modèles Excel ───────────────────────────────────────────────

_ACCENT = '1a1a1a'
_ACCENT_LIGHT = 'f0f0f0'
_RED = 'e8001d'


def _style_header(cell, required=False):
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(
        start_color=_RED if required else _ACCENT,
        end_color=_RED if required else _ACCENT,
        fill_type='solid',
    )
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_example(cell):
    cell.fill = PatternFill(start_color=_ACCENT_LIGHT, end_color=_ACCENT_LIGHT, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')


def generate_template(model_type='product'):
    """
    Génère un fichier XLSX modèle avec les en-têtes et une ligne d'exemple.
    Retourne les bytes du fichier.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    if model_type == 'product':
        columns, example, title = PRODUCT_COLUMNS, PRODUCT_EXAMPLE, 'Produits'
    elif model_type == 'prebuilt':
        columns, example, title = PREBUILT_COLUMNS, PREBUILT_EXAMPLE, 'PC Pré-montés'
    elif model_type == 'onlyone':
        columns, example, title = ONLYONE_COLUMNS, ONLYONE_EXAMPLE, 'Only One PCs'
    else:
        raise ValueError(f'model_type inconnu : {model_type}')

    ws.title = title

    # Ligne 1 : en-têtes
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22

    for col_idx, (col_name, _, _, _, required) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _style_header(cell, required=required)
        width = max(len(col_name) + 4, 16)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 30)

    # Ligne 2 : exemple
    for col_idx, (col_name, _, _, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=example.get(col_name, ''))
        _style_example(cell)

    # Feuille d'aide
    ws_help = wb.create_sheet('Aide — Colonnes')
    ws_help.column_dimensions['A'].width = 28
    ws_help.column_dimensions['B'].width = 55
    ws_help.column_dimensions['C'].width = 12

    help_headers = ['Colonne', 'Description', 'Requis']
    for col_idx, h in enumerate(help_headers, 1):
        cell = ws_help.cell(row=1, column=col_idx, value=h)
        _style_header(cell, required=False)
    ws_help.row_dimensions[1].height = 24

    for row_idx, (col_name, _, _, desc, required) in enumerate(columns, 2):
        ws_help.cell(row=row_idx, column=1, value=col_name)
        ws_help.cell(row=row_idx, column=2, value=desc)
        ws_help.cell(row=row_idx, column=3, value='✅' if required else '')
        if row_idx % 2 == 0:
            for c in range(1, 4):
                ws_help.cell(row=row_idx, column=c).fill = PatternFill(
                    start_color='f8f8f8', end_color='f8f8f8', fill_type='solid'
                )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
